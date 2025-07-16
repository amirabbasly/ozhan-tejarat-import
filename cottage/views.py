from rest_framework import viewsets, status, filters
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action
from .models import Cottage, CottageGoods, ExportedCottages, Expenses
from decimal import Decimal  # Import Decimal
from proforma.models import Performa  # Import Performa model
from django.http import HttpResponse, JsonResponse
from urllib.parse import urljoin
from django.views.decorators.http import require_http_methods
from django.middleware.csrf import get_token
from .serializers import CottageSerializer,CottageGoodsSerializer ,CustomsDeclarationInputSerializer, GreenCustomsDeclarationInputSerializer, CottageSaveSerializer, ExportedCottagesSerializer, FetchCotageRemainAmountSerializer, CottageNumberSerializer, ExpensesSerializer
import requests
import logging
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view
from django.core.files.storage import default_storage
from django.utils.text import slugify
import os
import uuid
from django.db import transaction
from accounts.permissions import IsAdmin, IsEditor , IsViewer
from collections import defaultdict
from .utils import get_cottage_combined_data
from openai import OpenAI
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from .filters import CottageFilter, ExpensesFilter, CottageGoodsFilter
from django.db.models import Q
import re
from openpyxl import Workbook
import datetime







class CustomPageNumberPagination(PageNumberPagination):
    page_size_query_param = 'page_size'
    max_page_size = 600

API_KEY = "AIzaSyBXmy7WymgH6np_beLSTR4MPASp23DBapw"
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={}".format(API_KEY)

class ChatbotAPIView(APIView):
    def post(self, request, *args, **kwargs):
        user_input = request.data.get('message', '')
        if not user_input:
            return Response({"error": "Message is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Extract cottage number(s) from user_input
        numbers = re.findall(r'\b\d+\b', user_input)  # e.g. "35411581"
        
        if numbers:
            cottage_number = numbers[0]
            cottages = Cottage.objects.filter(
                Q(cottage_number__icontains=cottage_number)
            )
        else:
            # fallback if no numbers found in user input
            cottages = Cottage.objects.none()

        # Build the context string from the data
        data_snippets = []
        for cottage in cottages:
            data_snippets.append(
                f"""
                number: {cottage.cottage_number}
                date: {cottage.cottage_date}
                proforma: {cottage.proforma}
                reference_number: {cottage.refrence_number}
                total_value: {cottage.total_value}
                quantity: {cottage.quantity}
                currency_price: {cottage.currency_price}
                cottage_customer: {cottage.cottage_customer}
                cottage_status: {cottage.cottage_status}
                rafee_taahod: {cottage.rafee_taahod}
                docs_recieved: {cottage.docs_recieved}
                rewatch: {cottage.rewatch}
                documents: {cottage.documents}
                """
            )

        database_context = "\n\n".join(data_snippets)

        if not database_context:
            database_context = "No relevant data found in the database."

        # Construct a prompt for Gemini
        prompt_text = f"""
        You are a chatbot with access to the following database records:

        {database_context}

        User's Question: {user_input}

        Please answer based on the above database information.
        """
        print(prompt_text)

        data = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": prompt_text
                        }
                    ]
                }
            ]
        }

        # Send the request to Gemini
        try:
            response = requests.post(
                GEMINI_API_URL,
                json=data,
                headers={"Content-Type": "application/json"},
                verify=False
            )

            if response.status_code == 200:
                try:
                    response_data = response.json()
                    print("Gemini API Response:", response_data)  # debug

                    candidates = response_data.get("candidates", [])
                    if candidates:
                        parts = candidates[0].get("content", {}).get("parts", [])
                        if parts:
                            bot_reply = parts[0].get("text", "").strip()
                            if bot_reply:
                                return Response({"reply": bot_reply}, status=status.HTTP_200_OK)
                            else:
                                return Response(
                                    {"error": "Received empty reply from Gemini."},
                                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                                )
                        else:
                            return Response(
                                {"error": "No content parts in the response."},
                                status=status.HTTP_500_INTERNAL_SERVER_ERROR
                            )
                    else:
                        return Response(
                            {"error": "No candidates returned from Gemini."},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR
                        )
                except requests.exceptions.JSONDecodeError:
                    return Response(
                        {"error": "Failed to decode Gemini API response."},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            else:
                return Response(
                    {"error": f"Request failed with status code {response.status_code}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        except requests.exceptions.RequestException as e:
            return Response({"error": f"Request failed: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class CottageCombinedDataView(APIView):
    def get(self, request):
        # Get the selected year from query params
        selected_year = request.query_params.get('year')
        if selected_year:
            # Validate the year if provided
            if not selected_year.isdigit():
                raise ValidationError({"error": "A valid year must be provided."})
            selected_year = int(selected_year)

        # Get combined data (with or without monthly data)
        data = get_cottage_combined_data(selected_year)
        return Response(data)


class FetchGoodsAPIView(APIView):
    permission_classes = [IsAdmin]
    def post(self, request):
        """
        Fetch goods for a given declaration using the external API.
        """
        gcuVcodeInt = request.data.get('gcuVcodeInt')
        ssdsshGUID = request.data.get('ssdsshGUID')
        urlVCodeInt = request.data.get('urlVCodeInt')

        if not gcuVcodeInt or not ssdsshGUID or not urlVCodeInt:
            return Response({'error': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

        # Payload for the external API
        payload = {
            'gcuVcodeint': gcuVcodeInt,
            'ssdsshGUID': ssdsshGUID,
            'urlVCodeInt': urlVCodeInt,
        }

        try:
            # Call the external API
            external_api_url = 'https://www.ntsw.ir/users/Ac/Gateway/FacadeRest/api/Customs/NTSW_GetAllGreenCustomsDeclarationGoods_V2'
            external_response = requests.post(
                external_api_url,
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=30
            )
            external_response.raise_for_status()

            # Parse and return the external API response
            data = external_response.json()
            if data.get('ErrorCode') != 0:
                logger.error(f"External API Error: {data.get('ErrorDesc')}")
                return Response({'error': data.get('ErrorDesc', 'Error fetching goods')}, status=status.HTTP_400_BAD_REQUEST)

            return Response(data, status=status.HTTP_200_OK)

        except requests.exceptions.RequestException as e:
            logger.error(f"Error calling external API: {e}")
            return Response({'error': 'Error fetching goods from external API'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SaveCottageView(APIView):
    permission_classes = [IsAdmin]
    serializer_class = CottageSaveSerializer

    def post(self, request):
        data = request.data
        cottage_number = data.get('cottage_number')

        try:
            # Attempt to retrieve the existing Cottage
            instance = Cottage.objects.get(cottage_number=cottage_number)
            serializer = CottageSaveSerializer(instance, data=data)
            is_update = True
        except Cottage.DoesNotExist:
            # Cottage does not exist; proceed to create a new one
            serializer = CottageSaveSerializer(data=data)
            is_update = False

        if not serializer.is_valid():
            logger.debug(f"Invalid data in SaveCottageView: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Save the Cottage instance (create or update)
                cottage = serializer.save()

                # Handle CottageGoods
                goods_data = data.get('goods', [])
                for good in goods_data:
                    CottageGoods.objects.update_or_create(
                        goodscode=good.get('goodscode'),
                        cottage=cottage,
                        defaults={
                            'customs_value': Decimal(good.get('customs_value', '0')),
                            'import_rights': Decimal(good.get('import_rights', '0')),
                            'red_cersent': Decimal(good.get('red_cersent', '0')),
                            'total_value': Decimal(good.get('total_value', '0')),
                            'added_value': Decimal(good.get('added_value', '0')),
                            'discount': Decimal(good.get('discount', '0')),
                            'quantity': good.get('quantity', 0),
                            'goods_description': good.get('goods_description', ''),
                            # Add other fields as necessary
                        }
                    )

            if is_update:
                return Response({'message': 'Cottage updated successfully!'}, status=status.HTTP_200_OK)
            else:
                return Response({'message': 'Cottage created successfully!'}, status=status.HTTP_201_CREATED)

        except Performa.DoesNotExist:
            return Response({'error': 'Proforma not found.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error in SaveCottageView: {e}", exc_info=True)
            return Response({'error': 'An unexpected error occurred.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SaveCottageGoodsView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        data = request.data
        cottage_number = data.get("cottage_number")

        try:
            cottage = Cottage.objects.get(cottage_number=cottage_number)
        except Cottage.DoesNotExist:
            return Response(
                {"error": "Cottage not found."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        goods_data = data.get("goods", [])

        # ---------- 1. build a set of the *new* codes -----------------------
        new_codes = set()
        for good in goods_data:
            code = good.get("ggsVcodeInt") or good.get("goodscode")
            if code:                      # skip rows with no key
                new_codes.add(code)
                CottageGoods.objects.update_or_create(
                    cottage=cottage,
                    goodscode=code,
                    defaults={
                        "customs_value":  Decimal(good.get("customs_value",  "0")),
                        "import_rights":  Decimal(good.get("import_rights", "0")),
                        "red_cersent":    Decimal(good.get("red_cersent",   "0")),
                        "total_value":    Decimal(good.get("total_value",   "0")),
                        "added_value":    Decimal(good.get("added_value",   "0")),
                        "discount":       Decimal(good.get("discount",      "0")),
                        "quantity":       Decimal(good.get("quantity",      "0")),
                        "goods_description":
                            good.get("goods_description")
                            or good.get("ggscommodityDescription", ""),
                    },
                )

        # ---------- 2. purge codes that disappeared from the payload -------
        (
            CottageGoods.objects
            .filter(cottage=cottage)
            .exclude(goodscode__in=new_codes)
            .delete()
        )
        # -------------------------------------------------------------------

        return Response(
            {"message": "Goods saved successfully!"},
            status=status.HTTP_201_CREATED,
        )

class FetchCustomsDutyInformationAPIView(APIView):
    permission_classes = [IsAdmin]
    """
    Fetch customs duty information for a specific goods item using the external API.
    """
    def post(self, request):
        ggsVcodeInt = request.data.get('ggsVcodeInt')
        ssdsshGUID = request.data.get('ssdsshGUID')
        urlVCodeInt = request.data.get('urlVCodeInt')

        if not ggsVcodeInt or not ssdsshGUID or not urlVCodeInt:
            return Response({'error': 'Missing required parameters'}, status=status.HTTP_400_BAD_REQUEST)

        payload = {
            'ggsVcodeInt': ggsVcodeInt,
            'ssdsshGUID': ssdsshGUID,
            'urlVCodeInt': urlVCodeInt,
        }

        try:
            # Call the external API
            external_api_url = 'https://www.ntsw.ir/users/Ac/Gateway/FacadeRest/api/Customs/NTSW_GetAllGreenCustomsDutyInformation_V2'
            external_response = requests.post(
                external_api_url,
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=30
            )
            external_response.raise_for_status()

            data = external_response.json()
            if data.get('ErrorCode') != 0:
                logger.error(f"External API Error: {data.get('ErrorDesc')}")
                return Response({'error': data.get('ErrorDesc', 'Error fetching customs duty information')}, status=status.HTTP_400_BAD_REQUEST)

            return Response(data, status=status.HTTP_200_OK)

        except requests.exceptions.RequestException as e:
            logger.error(f"Error calling external API: {e}")
            return Response({'error': 'Error fetching customs duty information from external API'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CottageViewSet(viewsets.ModelViewSet):
    queryset = Cottage.objects.all()
    serializer_class = CottageSerializer
    pagination_class = CustomPageNumberPagination  # Apply pagination here
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = CottageFilter
    search_fields = ["total_value","customs_value","cottage_number", "proforma__prf_order_no",]  # fields you want to search

    def get_serializer_class(self):
        if self.action == 'post':
            return CottageSaveSerializer
        elif self.action == 'create':
            return CottageSaveSerializer
        return super().get_serializer_class()

    # Custom action to get a cottage by number (read-only)
    @action(detail=False, methods=['get'], url_path='by-number/(?P<cottage_number>[^/.]+)')
    def get_by_cottage_number(self, request, cottage_number=None):
        try:
            cottage = Cottage.objects.get(cottage_number=cottage_number)
            serializer = CottageSerializer(cottage)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Cottage.DoesNotExist:
            return Response(
                {"error": "Cottage not found."},
                status=status.HTTP_404_NOT_FOUND
            )
    @action(detail=False, methods=["get"], url_path="numbers")
    def list_numbers(self, request):
        """
        /api/cottages/numbers/ → [{id:…, cottage_number:…}, …]
        """
        qs = self.get_queryset()
        page = self.paginate_queryset(qs)
        serializer = CottageNumberSerializer(page or qs, many=True)
        return self.get_paginated_response(serializer.data) if page else Response(serializer.data)


    # New custom action to delete selected cottages
    @action(detail=False, methods=['post'], url_path='delete-selected')
    def delete_selected(self, request):
        """
        Custom action to delete selected cottages.
        Expects a list of cottage IDs in the request data.
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'No IDs provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate IDs
        if not isinstance(ids, list):
            return Response({'error': 'IDs should be provided as a list.'}, status=status.HTTP_400_BAD_REQUEST)

        # Optionally, validate that the IDs are integers
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return Response({'error': 'Invalid IDs. IDs must be integers.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cottages = Cottage.objects.filter(id__in=ids)
            if not cottages.exists():
                return Response({'error': 'No cottages found for the provided IDs.'}, status=status.HTTP_404_NOT_FOUND)
            deleted_count = cottages.count()
            cottages.delete()
            return Response({'message': f'{deleted_count} cottages deleted.'}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error deleting cottages: {e}", exc_info=True)
            return Response({'error': 'An error occurred while deleting cottages.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
logger = logging.getLogger(__name__)
class CustomsDeclarationListView(APIView):
    permission_classes = [IsAdmin]
    def post(self, request):
        # Initialize the serializer with the incoming data
        serializer = CustomsDeclarationInputSerializer(data=request.data)
        if not serializer.is_valid():
            # If validation fails, return errors to the user
            logger.debug(f"Invalid input data: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Extract validated data
        validated_data = serializer.validated_data
        ssdsshGUID = str(validated_data['ssdsshGUID'])  # Convert UUID to string if necessary
        urlVCodeInt = validated_data['urlVCodeInt']
        page_size = validated_data['PageSize']
        start_index = validated_data['StartIndex']
        
        # Build the payload with user-provided values
        payload = {
            "Count": 0,
            "CountryID": 0,
            "DeclarationType": 0,
            "EntranceCustomsName": "",
            "FromDeclarationDate": None,
            "FullSerialNumber": "",
            "NationalCode": "",
            "OrderRegistrationNumber": "",
            "PageSize": page_size,  # User-provided
            "PateNumber": "",
            "StartIndex": start_index, # User-provided
            "ToDeclarationDate": None,
            "gcudeclarationStatus": "",
            "gculCReferenceNumber": "",
            "ssdsshGUID": ssdsshGUID,  # User-provided
            "urlVCodeInt": urlVCodeInt  # User-provided
        }

        # Add additional filters from the request data if necessary
        # Example:
        # if 'someFilter' in request.data:
        #     payload['someFilter'] = request.data['someFilter']

        # Log payload for debugging
        logger.debug(f"Payload sent to NTSW API: {payload}")

        try:
            # Make the API call
            response = requests.post(
                'https://www.ntsw.ir/users/Ac/Gateway/FacadeRest/api/Customs/NTSW_GetAllCustomizeCustomsDeclaration',
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=30
            )
            response.raise_for_status()

            # Parse the response
            data = response.json()
            logger.debug(f"Response received from NTSW API: {data}")

            if data.get('ErrorCode') != 0:
                error_desc = data.get('ErrorDesc', 'Unknown error occurred.')
                logger.error(f"NTSW API Error: {error_desc}")
                return Response({'error': error_desc}, status=status.HTTP_400_BAD_REQUEST)

            return Response(data, status=status.HTTP_200_OK)

        except requests.exceptions.RequestException as e:
            logger.error(f"RequestException: {e}")
            return Response({'error': 'Error fetching data from NTSW API'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            logger.error(f"Unexpected Error: {e}", exc_info=True)
            return Response({'error': 'An unexpected error occurred on the server.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['POST'])
def upload_file(request, cottage_id):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        
        # Get the original filename and file extension
        original_filename = file.name
        file_extension = os.path.splitext(original_filename)[1]  # Extract the file extension
        
        # Ensure the file name is safe (no path traversal)
        safe_filename = original_filename.replace(" ", "_")  # Replace spaces with underscores, if necessary

        # Construct the file path within the cottage directory
        cottage_directory = os.path.join('uploads', 'cottages', str(cottage_id))
        file_path = os.path.join(cottage_directory, safe_filename)

        # Save the file using Django's default storage
        file_path = default_storage.save(file_path, file)

        # Associate the file with the cottage instance
        try:
            cottage = Cottage.objects.get(id=cottage_id)
            cottage.documents = file_path  # Set the file path
            cottage.save()

            return JsonResponse({'file_path': file_path}, status=status.HTTP_201_CREATED)
        except Cottage.DoesNotExist:
            return JsonResponse({'error': 'Cottage not found.'}, status=status.HTTP_404_NOT_FOUND)
class GreenCustomsDeclarationView(APIView):
    """
    API view to handle requests for Green Customs Declarations.
    """
    permission_classes = [IsAdmin]
    def post(self, request):
        # Validate incoming data
        serializer = GreenCustomsDeclarationInputSerializer(data=request.data)
        if not serializer.is_valid():
            logger.debug(f"Invalid input data: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Extract validated data
        validated_data = serializer.validated_data

        FullSerilaNumber = validated_data['FullSerilaNumber']
        ssdsshGUID = str(validated_data['ssdsshGUID'])  # Convert UUID to string
        urlVCodeInt = validated_data['urlVCodeInt']

        # Build the payload for the external API
        payload = {
            "DeclarationType":0,
            "FullSerilaNumber": FullSerilaNumber,
            "ssdsshGUID": ssdsshGUID,
            "urlVCodeInt": urlVCodeInt
        }

        logger.debug(f"Payload sent to NTSW_GetGreenCustomsDeclaration API: {payload}")

        # External API endpoint
        api_url = "https://www.ntsw.ir/users/Ac/Gateway/FacadeRest/api/Customs/NTSW_GetGreenCustomsDeclaration"

        headers = {
            "Content-Type": "application/json",
            # Add other headers if required, such as Authorization
        }

        try:
            # Make the POST request to the external API
            response = requests.post(
                api_url,
                json=payload,
                headers=headers,
                timeout=30  # seconds
            )
            response.raise_for_status()  # Raise HTTPError for bad responses

            # Parse the JSON response
            api_response = response.json()
            logger.debug(f"Response received from NTSW API: {api_response}")

            # Check for API-specific error codes
            if api_response.get('ErrorCode') != 0:
                error_desc = api_response.get('ErrorDesc', 'Unknown error occurred.')
                logger.error(f"NTSW API Error: {error_desc}")
                return Response({'error': error_desc}, status=status.HTTP_400_BAD_REQUEST)

            # Optionally, you can serialize the response data here
            # For simplicity, we'll return the response as-is
            return Response(api_response, status=status.HTTP_200_OK)

        except requests.exceptions.Timeout:
            logger.error("The request to NTSW API timed out.")
            return Response({'error': 'The request to the external API timed out.'}, status=status.HTTP_504_GATEWAY_TIMEOUT)
        except requests.exceptions.ConnectionError:
            logger.error("There was a connection error while contacting NTSW API.")
            return Response({'error': 'Connection error with external API.'}, status=status.HTTP_502_BAD_GATEWAY)
        except requests.exceptions.HTTPError as http_err:
            logger.error(f"HTTP error occurred: {http_err}")
            return Response({'error': 'HTTP error occurred when contacting external API.'}, status=status.HTTP_502_BAD_GATEWAY)
        except requests.exceptions.RequestException as req_err:
            logger.error(f"RequestException: {req_err}")
            return Response({'error': 'An error occurred while requesting external API.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except ValueError:
            logger.error("Invalid JSON received from NTSW API.")
            return Response({'error': 'Invalid response format from external API.'}, status=status.HTTP_502_BAD_GATEWAY)
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            return Response({'error': 'An unexpected error occurred on the server.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CottageGoodsViewSet(viewsets.ModelViewSet):
    queryset = CottageGoods.objects.all()
    serializer_class = CottageGoodsSerializer
    pagination_class = CustomPageNumberPagination  # Apply pagination here
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = CottageGoodsFilter
    search_fields = ["cottage__cottage_number", ]  # fields you want to search

    # Add custom logic or actions if needed
    def perform_create(self, serializer):
        # Example: Add additional validations or default values
        serializer.save()
class ExportCustomsDeclarationListView(APIView):
    permission_classes = [IsAdmin]
    def post(self, request):
        # Initialize the serializer with the incoming data
        serializer = CustomsDeclarationInputSerializer(data=request.data)
        if not serializer.is_valid():
            # If validation fails, return errors to the user
            logger.debug(f"Invalid input data: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Extract validated data
        validated_data = serializer.validated_data
        ssdsshGUID = str(validated_data['ssdsshGUID'])  # Convert UUID to string if necessary
        urlVCodeInt = validated_data['urlVCodeInt']
        page_size = validated_data['PageSize']
        start_index = validated_data['StartIndex']
        
        # Build the payload with user-provided values
        payload = {
            "Count": 0,
            "DeclarationType": 1,
            "FullSerialNumber": "",
            "NationalCode": "",
            "OrderRegistrationNumber": "",
            "PageSize": page_size,  # User-provided
            "PateNumber": "",
            "StartIndex": start_index, # User-provided
            "gcudeclarationStatus": "",
            "gculCReferenceNumber": "",
            "ssdsshGUID": ssdsshGUID,  # User-provided
            "urlVCodeInt": urlVCodeInt  # User-provided
        }

        # Log payload for debugging
        logger.debug(f"Payload sent to NTSW API: {payload}")

        try:
            # Make the API call
            response = requests.post(
                'https://www.ntsw.ir/users/Ac/Gateway/FacadeRest/api/Export/GetAllCustomizeCustomsDeclaration',
                json=payload,
                headers={'Content-Type': 'application/json'},
                timeout=30
            )
            response.raise_for_status()

            # Parse the response
            data = response.json()
            logger.debug(f"Response received from NTSW API: {data}")

            if data.get('ErrorCode') != 0:
                error_desc = data.get('ErrorDesc', 'Unknown error occurred.')
                logger.error(f"NTSW API Error: {error_desc}")
                return Response({'error': error_desc}, status=status.HTTP_400_BAD_REQUEST)

            return Response(data, status=status.HTTP_200_OK)

        except requests.exceptions.RequestException as e:
            logger.error(f"RequestException: {e}")
            return Response({'error': 'Error fetching data from NTSW API'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            logger.error(f"Unexpected Error: {e}", exc_info=True)
            return Response({'error': 'An unexpected error occurred on the server.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class FetchCotageRemainAmountView(APIView):
    """
    API View to fetch CotageRemainAmount from the second external API and save/update ExportedCottages.
    """
    permission_classes = [IsAdmin]  # Restrict access to admin users; adjust as needed

    def post(self, request):
        # 1. Validate Input Data
        serializer = FetchCotageRemainAmountSerializer(data=request.data)
        if not serializer.is_valid():
            logger.debug(f"Invalid input data: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        ssdsshGUID = validated_data['ssdsshGUID']
        urlVcodeInt = validated_data['urlVcodeInt']
        full_serial_number = validated_data['fullSerialNumber']
        total_value = validated_data['total_value']
        cottage_date = validated_data['cottage_date']
        quantity = validated_data['quantity']
        currency_type = validated_data['currency_type']
        declaration_status = validated_data['status']
        
        # 2. Split FullSerialNumber into CustomCode and CotageCode
        try:
            custom_code, cotage_code = full_serial_number.split('-')
        except ValueError:
            logger.warning(f"Invalid FullSerialNumber format: {full_serial_number}")
            return Response({
                "error": "Invalid FullSerialNumber format. Expected 'CustomCode-CotageCode'."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 3. Prepare Payload for Second API
        second_api_payload = {
            "CotageCode": cotage_code,
            "CustomCode": custom_code,
            "InquiryType": 0,
            "OriginalCaller": 0,
            "SessionID": str(ssdsshGUID),  
            "nationalcode": None,
            "urlVCodeInt": urlVcodeInt
        }
        
        logger.debug(f"Sending payload to second API: {second_api_payload}")
        
        # 4. Call the Second External API
        try:
            second_api_response = requests.post(
                "https://www.ntsw.ir/users/ac/gateway/facaderest/api/ObligationEliminate/CotageInquiry",
                json=second_api_payload,
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            second_api_response.raise_for_status()
            second_data = second_api_response.json()
            logger.debug(f"Response from second API: {second_data}")
        except requests.RequestException as e:
            logger.error(f"Second API RequestException for FullSerialNumber={full_serial_number}: {e}")
            return Response({
                "error": "Error calling the second API."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except ValueError as e:
            # JSON decoding failed
            logger.error(f"JSON decode error from second API for FullSerialNumber={full_serial_number}: {e}")
            return Response({
                "error": "Invalid JSON response from the second API."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            logger.error(f"Unexpected error calling second API for FullSerialNumber={full_serial_number}: {e}", exc_info=True)
            return Response({
                "error": "Unexpected server error calling the second API."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # 5. Check for Errors in Second API Response
        if second_data.get("ErrorCode") != 0:
            error_desc = second_data.get("ErrorDesc", "Unknown error from second API.")
            logger.error(f"Second API Error for FullSerialNumber={full_serial_number}: {error_desc}")
            return Response({
                "error": error_desc
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # 6. Extract CotageRemainAmount
        cotage_info = second_data.get("CotageInformation", {})
        cotage_remain_amount = cotage_info.get("CotageRemainAmount")
        
        if cotage_remain_amount is None:
            logger.warning(f"CotageRemainAmount not found for FullSerialNumber: {full_serial_number}")
            return Response({
                "error": "CotageRemainAmount not found in second API response."
            }, status=status.HTTP_404_NOT_FOUND)
        try:
           cotage_remain_dec = Decimal(cotage_remain_amount)
        except (ValueError, TypeError, InvalidOperation):
            logger.error(f"Invalid CotageRemainAmount format: {cotage_remain_amount}")
            return Response({"error": "Invalid CotageRemainAmount format."},
                    status=status.HTTP_400_BAD_REQUEST) 
        # 7. Save or Update Data to ExportedCottages Model
        exported_cottage, created = ExportedCottages.objects.update_or_create(
            full_serial_number=full_serial_number,
            defaults={

                "remaining_total": cotage_remain_dec,
                "declaration_status": declaration_status,
                "currency_type": currency_type,
                "total_value": total_value,
                "cottage_date": cottage_date,
                "quantity": quantity,
            }
        )
        
        if created:
            logger.info(f"ExportedCottages entry created: {exported_cottage}")
            message = "ExportedCottages entry successfully created."
            status_code = status.HTTP_201_CREATED
        else:
            logger.info(f"ExportedCottages entry updated: {exported_cottage}")
            message = "ExportedCottages entry successfully updated."
            status_code = status.HTTP_200_OK
        
        # 8. Construct Response to Client
        response_data = {
            "id": exported_cottage.id,
            "full_serial_number": exported_cottage.full_serial_number,
            "remaining_total": exported_cottage.remaining_total,
            "quantity": exported_cottage.quantity,
            "message": message
        }
        
        return Response(response_data, status=status_code)

class ExportedCottagesViewSet(viewsets.ModelViewSet):
    queryset = ExportedCottages.objects.all()
    serializer_class = ExportedCottagesSerializer

    # Add custom logic or actions if needed
    def perform_create(self, serializer):
        serializer.save()

    # New custom action to delete selected cottages
    @action(detail=False, methods=['post'], url_path='delete-selected')
    def delete_selected(self, request):
        """
        Custom action to delete selected cottages.
        Expects a list of cottage IDs in the request data.
        """
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'No IDs provided.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate IDs
        if not isinstance(ids, list):
            return Response({'error': 'IDs should be provided as a list.'}, status=status.HTTP_400_BAD_REQUEST)

        # Optionally, validate that the IDs are integers
        try:
            ids = [int(id) for id in ids]
        except ValueError:
            return Response({'error': 'Invalid IDs. IDs must be integers.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cottages = ExportedCottages.objects.filter(id__in=ids)
            if not cottages.exists():
                return Response({'error': 'No cottages found for the provided IDs.'}, status=status.HTTP_404_NOT_FOUND)
            deleted_count = cottages.count()
            cottages.delete()
            return Response({'message': f'{deleted_count} cottages deleted.'}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error deleting cottages: {e}", exc_info=True)
            return Response({'error': 'An error occurred while deleting cottages.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        # Custom action to get a cottage by number (read-only)
    @action(detail=False, methods=['get'], url_path='by-number/(?P<full_serial_number>[^/.]+)')
    def get_by_full_serial_number(self, request, full_serial_number=None):
        try:
            cottage = ExportedCottages.objects.get(full_serial_number=full_serial_number)
            serializer = ExportedCottagesSerializer(cottage)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Cottage.DoesNotExist:
            return Response(
                {"error": "Cottage not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class ImportCottagesView(APIView):
    """
    DRF view to handle Excel file uploads with Persian headings.
    """
    parser_classes = (MultiPartParser, FormParser,)  # allow file uploads

    EXPECTED_HEADINGS = {
        'سریال اظهارنامه': 'cottage_number',
        'تاریخ': 'cottage_date',
        'شماره ثبت سفارش': 'proforma',
        'کد ساتا': 'refrence_number',
        'ارزش کل': 'total_value',
        'تعداد': 'quantity',
        'مشتری': 'cottage_customer',
        'قیمت ارز': 'currency_price',
        'وضعیت اظهار': 'cottage_status',
    }

    def post(self, request, format=None):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "یک فایل انتخاب کنید."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active  # or workbook["SheetName"] if needed

            # Validate heading row
            heading_row = [cell.value for cell in sheet[1]]  # First row as list
            if len(heading_row) < len(self.EXPECTED_HEADINGS):
                return Response({"error": "تعداد ستون های فایل کامل نمیباشد"}, status=400)

            col_map = {}
            for col_index, heading in enumerate(heading_row):
                if heading in self.EXPECTED_HEADINGS:
                    col_map[col_index] = self.EXPECTED_HEADINGS[heading]

            if len(col_map) < len(self.EXPECTED_HEADINGS):
                missing_headings = set(self.EXPECTED_HEADINGS.keys()) - set(heading_row)
                return Response({"error": f"فایل باید شامل فیلد های: {', '.join(missing_headings)} باشد"}, status=400)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(value is None for value in row):  # Skip empty rows
                    continue

                data_for_model = {}
                for col_index, cell_value in enumerate(row):
                    if col_index in col_map:
                        field_name = col_map[col_index]
                        data_for_model[field_name] = cell_value

                # Handle the 'proforma' field (ForeignKey)
                proforma_value = data_for_model.get('proforma')
                if proforma_value:
                    try:
                        proforma_obj = Performa.objects.get(prf_order_no=proforma_value)
                        data_for_model['proforma'] = proforma_obj
                    except Performa.DoesNotExist:
                        return Response({"error": f"ثبت سفارش با شماره ثبت سفارش '{proforma_value}' وجود ندارد."}, status=400)

                cottage_number = data_for_model.get('cottage_number')
                if not cottage_number:
                    continue

                Cottage.objects.update_or_create(
                    cottage_number=cottage_number,
                    defaults=data_for_model
                )

            return Response({"success": "اظهانمامه ها با موفقیت بارگزاری شدند."}, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
class ImportExportedCottagesView(APIView):
    """
    DRF view to handle Excel file uploads with Persian headings.
    """
    parser_classes = (MultiPartParser, FormParser,)  # allow file uploads

    # We define the exact Persian headings we expect in row 1 -> map them to model fields
    EXPECTED_HEADINGS = {
        'شماره سریال کامل': 'full_serial_number',
        'تاریخ': 'cottage_date',
        'ارزش کل': 'total_value',
        'تعداد': 'quantity',
        'نوع ارز': 'currency_type',
        'قیمت ارز': 'currency_price',
        'وضعیت اظهار': 'declaration_status',
        'مقدار باقی‌مانده': 'remaining_total',
    }

    def post(self, request, format=None):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active  # or workbook["SheetName"] if needed

            # 1) Validate heading row
            heading_row = [cell.value for cell in sheet[1]]  # First row as list
            if len(heading_row) < len(self.EXPECTED_HEADINGS):
                return Response({"error": "تعداد ستون های فایل کامل نمیباشد"}, status=400)

            # Build a map from column index -> model field name
            # Example: col_map[0] = 'full_serial_number'
            col_map = {}
            for col_index, heading in enumerate(heading_row):
                if heading in self.EXPECTED_HEADINGS:
                    col_map[col_index] = self.EXPECTED_HEADINGS[heading]

            # Check we found all required headings
            if len(col_map) < len(self.EXPECTED_HEADINGS):
                missing_headings = set(self.EXPECTED_HEADINGS.keys()) - set(heading_row)
                return Response({
                    "error": f"فایل باید شامل فیلد های: {', '.join(missing_headings)} باشد"
                }, status=400)

            # 2) Iterate from row 2 onward
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows (all None)
                if all(value is None for value in row):
                    continue

                # Prepare a dict to hold the row's data for model fields
                data_for_model = {}

                # For each column in our row, see if col_map has a field name
                for col_index, cell_value in enumerate(row):
                    if col_index in col_map:
                        field_name = col_map[col_index]
                        data_for_model[field_name] = cell_value

                # If 'full_serial_number' is the unique key, we can update_or_create
                full_serial = data_for_model.get('full_serial_number')
                if not full_serial:
                    # If the row doesn't have a full_serial_number, skip or handle error
                    continue

                ExportedCottages.objects.update_or_create(
                    full_serial_number=full_serial,
                    defaults=data_for_model
                )

            return Response({"success": "اظهانمامه ها با موفقیت بارگزاری شدند."})

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class CottageExcelExportView(APIView):

    def post(self, request, *args, **kwargs):
        cottage_numbers = request.data.get('cottage_numbers', [])
        if not isinstance(cottage_numbers, list) or not cottage_numbers:
            return Response(
                {"detail": "Provide a non-empty list of cottage_numbers."},
                status=status.HTTP_400_BAD_REQUEST
            )

        qs = Cottage.objects.filter(cottage_number__in=cottage_numbers)

        # create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Cottages"

        # header row
        headers = [
            "شماره کوتاژ", "تاریخ",
            "شماره ثبت سفارش", 
            "شماره ساتا",
            "ارزش ارزی", "ارزش گمرکی",
             "مشتری",
            "وضعیت", 
             "ارزش افزوده",
             "بازبینی",
             "اخذ مدارک", 
              "رفع تعهد",                        
        ]
        ws.append(headers)

        # data rows
        for c in qs:
            rewatch_display = "بله" if getattr(c, "rewatch", False) else "خیر"
            docs_recieved_display = "بله" if getattr(c, "docs_recieved", False) else "خیر"
            rafee_taahod_display = "بله" if getattr(c, "rafee_taahod", False) else "خیر"
            ws.append([
                c.cottage_number,
                str(c.cottage_date),                
                c.proforma.prf_order_no,           
                c.refrence_number or "",
                float(c.total_value),
                float(c.customs_value or 0),
                (c.cottage_customer.full_name 
                   if c.cottage_customer else ""),
                c.cottage_status or "",
                float(c.added_value or 0),
                rewatch_display,
                docs_recieved_display,
                rafee_taahod_display

            ])

        # prepare HTTP response
        today = datetime.date.today().isoformat()
        filename = f"cottages_{today}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class ExpensesViewSet(viewsets.ModelViewSet):
    queryset = Expenses.objects.all()
    serializer_class = ExpensesSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ExpensesFilter
    pagination_class = CustomPageNumberPagination  # Apply pagination here
    search_fields = ["cottage__cottage_number","description","value" ]  # fields you want to search

class CottageGoodsExportView(APIView):
    def post( self , request, *args, **kwargs):
        goods_codes = request.data.get('goods_codes', [])
        if not isinstance(goods_codes, list) or not goods_codes:
            return Response(
                {"detail": "Provide a non-empty list of goodscodes."},
                status=status.HTTP_400_BAD_REQUEST
            )

        qs = CottageGoods.objects.filter(goodscode__in=goods_codes)
        # create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Cottage Goods"
        headers = [
            "شماره کوتاژ",
            "شماره ثبت سفارش", 
            "ارزش ارزی",
             "ارزش گمرکی",
             "حقوق ورودی",
            "ارزش افزوده", 
             "حلال احمر	",
        ]
        ws.append(headers)
        for c in qs:
            ws.append([
                c.cottage.cottage_number,
                c.cottage.proforma.prf_order_no,           
                float(c.total_value or 0),
                float(c.customs_value),
                float(c.added_value or 0),
                float(c.red_cersent or 0),


            ])
        today = datetime.date.today().isoformat()
        filename = f"cottage_goods_{today}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response