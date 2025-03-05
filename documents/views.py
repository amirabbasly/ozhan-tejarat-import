from django.conf import settings
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from PIL import Image, ImageDraw, ImageFont
import io
from .serializers import OverlayTextSerializer, ImageTemplateSerializer,  SellerSerializer, BuyerSerializer, InvoiceSerializer, OriginSerializer, ProformaInvoiceSerializer
from .models import ImageTemplate, Seller, Buyer, Invoice, ProformaInvoice
import os
import openpyxl
from reportlab.lib.pagesizes import letter
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
)

from rest_framework.decorators import action
from reportlab.graphics.shapes import Drawing, Line
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PyPDF2 import PdfMerger
from rest_framework.test import APIRequestFactory
from django.http import HttpResponse, HttpRequest

class OverlayTextView(APIView):
    def post(self, request):
        # 1) Fetch the template based on template_id
        try:
            template_id = request.data.get('template_id')
            template = ImageTemplate.objects.get(id=template_id)
        except ImageTemplate.DoesNotExist:
            return Response({"error": "Template not found."},
                            status=status.HTTP_404_NOT_FOUND)

        # 2) Validate incoming data - only invoice_id and template_id are needed
        serializer = OverlayTextSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors,
                            status=status.HTTP_400_BAD_REQUEST)

        # 3) Extract validated data from the serializer
        invoice_id = serializer.validated_data.get('invoice_id')

        # 4) Fetch the invoice from the database
        try:
            invoice = Invoice.objects.get(id=invoice_id)
        except Invoice.DoesNotExist:
            return Response({"error": "Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        # 5) Fetch related data from the models
        exporter = invoice.seller.seller_name + "\n" + invoice.seller.seller_address 
        consignee = invoice.buyer.buyer_name + "\n" + invoice.buyer.buyer_address  + "\n" + invoice.buyer.buyer_tel 
        means_of_transport = invoice.means_of_transport + " from " + invoice.port_of_loading + " to " + invoice.relevant_location
        number_of_invoices = invoice.invoice_number+ "\n" + str(invoice.invoice_date)
        goods = []

        # Collect goods data (items from the invoice)
        for item in invoice.items.all():
            goods.append({
                'description': str(item.pack) + " PCK"+"   "  + item.description,
                'hscode': item.commodity_code,
                'quantity': str(item.gw_kg) + "KG",  # Make sure it's a string
                'number_of_invoices': str(invoice.invoice_number) + "\n" + str(invoice.invoice_date)
            })

        # 6) Load the template image and prepare for drawing
        img = Image.open(template.template_image.path)
        draw = ImageDraw.Draw(img)
        font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial.ttf')
        font_size = template.font_size  # Read from the model
        font = ImageFont.truetype(font_path, font_size)

        # 7) Ensure positions are integers, using `map(int, ...)` for splitting
        exporter_x, exporter_y = map(int, template.exporter_position.split(','))
        consignee_x, consignee_y = map(int, template.consignee_position.split(','))
        transport_x, transport_y = map(int, template.means_of_transport_position.split(','))
        invoices_x, invoices_y = map(int, template.number_of_invoices_position.split(','))

        # Draw static fields (exporter, consignee, transport)
        draw.multiline_text((exporter_x, exporter_y), exporter, fill="black", font=font)
        draw.multiline_text((consignee_x, consignee_y), consignee, fill="black", font=font)
        draw.multiline_text((transport_x, transport_y), means_of_transport , fill="black", font=font)
        draw.multiline_text((invoices_x, invoices_y), number_of_invoices , fill="black", font=font)


        # 8) Use the same positions for seller_name and buyer_name (as they are the same as consignee and exporter)
        seller_name_x, seller_name_y = map(int, template.exporter_position.split(','))
        buyer_name_x, buyer_name_y = map(int, template.consignee_position.split(','))
        draw.multiline_text((seller_name_x, seller_name_y), invoice.seller.seller_name, fill="black", font=font)
        draw.multiline_text((buyer_name_x, buyer_name_y), invoice.buyer.buyer_name, fill="black", font=font)

        # 9) Draw each good in its own row (items)
        description_x = int(template.description_position)  # Just X coordinate as integer
        hscode_x = int(template.hscode_position)  # Just X coordinate as integer
        quantity_x = int(template.quantity_position)  # Just X coordinate as integer

        goods_start_x, goods_start_y = map(int, template.goods_start_position.split(','))
        current_y = goods_start_y
        line_height = template.goods_line_height  # Vertical spacing between each row
        index_x = goods_start_x

        for i, good in enumerate(goods, start=1):
            description_text = str(good['description'])  # Ensure it's a string
            hscode_text = str(good['hscode'])  # Ensure it's a string
            quantity_text = str(good['quantity'])  # Ensure it's a string
            invoices_text = str(good['number_of_invoices'])  # Ensure it's a string

            # Draw each field at its respective X position and the current Y position
            draw.multiline_text((index_x, current_y), str(i), fill="black", font=font)
            draw.multiline_text((description_x, current_y), description_text, fill="black", font=font)
            draw.multiline_text((hscode_x, current_y), hscode_text, fill="black", font=font)
            draw.multiline_text((quantity_x, current_y), quantity_text, fill="black", font=font)

            # Move down to the next row
            current_y += line_height

        img_io = BytesIO()
        img.save(img_io, format="PDF")  # Save the image as a PDF (instead of JPEG)
        img_io.seek(0)

        # Set the filename and content type for the response
        filename = f"certificate_of_origin_{invoice.id}.pdf"
        response = HttpResponse(img_io.getvalue(), content_type="application/pdf")
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

class TemplateListView(APIView):    
    def get(self, request):
        # Get all templates from the database
        templates = ImageTemplate.objects.all()
        serializer = ImageTemplateSerializer(templates, many=True)
        return Response(serializer.data)

class SellerViewSet(viewsets.ModelViewSet):
    queryset = Seller.objects.all()
    serializer_class = SellerSerializer
    pagination_class = None  # Disable pagination here
    @action(detail=False, methods=['get'], url_path='by-id/(?P<seller_id>[^/.]+)')
    def get_seller_by_id(self, request, seller_id=None):
        try:
            seller = Seller.objects.get(id=seller_id)
            serializer = SellerSerializer(seller)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Seller.DoesNotExist:
            return Response(
                {"error": "Seller not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class BuyerViewSet(viewsets.ModelViewSet):
    queryset = Buyer.objects.all()
    serializer_class = BuyerSerializer
    pagination_class = None  # Disable pagination here
    @action(detail=False, methods=['get'], url_path='by-id/(?P<buyer_id>[^/.]+)')
    def get_buyer_by_id(self, request, buyer_id=None):
        try:
            buyer = Buyer.objects.get(id=buyer_id)
            serializer = BuyerSerializer(buyer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Buyer.DoesNotExist:
            return Response(
                {"error": "Seller not found."},
                status=status.HTTP_404_NOT_FOUND
            )


class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

    @action(detail=False, methods=['get'], url_path='by-number/(?P<invoice_id>[^/.]+)')
    def get_by_invoice_number(self, request, invoice_id=None):
        try:
            invoice = Invoice.objects.get(invoice_id=invoice_id)
            serializer = InvoiceSerializer(invoice)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Invoice.DoesNotExist:
            return Response(
                {"error": "Invoice not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class ProformaInvoiceViewSet(viewsets.ModelViewSet):
    queryset = ProformaInvoice.objects.all()
    serializer_class = ProformaInvoiceSerializer

    @action(detail=False, methods=['get'], url_path='by-number/(?P<proforma_invoice_id>[^/.]+)')
    def get_by_invoice_number(self, request, proforma_invoice_id=None):
        try:
            proforma_invoice = ProformaInvoice.objects.get(proforma_invoice_id=proforma_invoice_id)
            serializer = ProformaInvoiceSerializer(proforma_invoice)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ProformaInvoice.DoesNotExist:
            return Response(
                {"error": "Invoice not found."},
                status=status.HTTP_404_NOT_FOUND
            )

class InvoicePDFView(APIView):
    """
    Returns a PDF with the invoice number and date on the right side,
    and other sections laid out as per the template.
    """
    def get(self, request, pk, format=None):
        try:
            invoice = Invoice.objects.get(pk=pk)
        except Invoice.DoesNotExist:
            return Response({"detail": "Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        # Register the Broadway font
        pdfmetrics.registerFont(TTFont('Broadway', 'static/fonts/broadway.ttf'))

        # 1) Create the HttpResponse for the PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{pk}.pdf"'

        # Set up the SimpleDocTemplate with custom margins (left, right, top, bottom)
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            title=f"Invoice {invoice.invoice_number}",
            author="Your Company",
            leftMargin=0.1 * inch,   # 0.5 inch left margin
            rightMargin=0.1 * inch,  # 0.5 inch right margin
            topMargin=0.2 * inch,    # 0.2 inch top margin
            bottomMargin=0.5 * inch  # 0.5 inch bottom margin
        )

        # 3) Get default styles and add custom styles
        styles = getSampleStyleSheet()

        styles.add(ParagraphStyle(
            name="InvoiceTitle",
            fontSize=24,
            textColor=colors.black,
            alignment=1,  # Center align
            spaceAfter=0
        ))

        styles.add(ParagraphStyle(
            name="SellerName",
            fontName="Broadway",
            fontSize=26,
            textColor=colors.red,
            alignment=1,  # Center align
            spaceAfter=20,
        ))

        styles.add(ParagraphStyle(
            name="RightAlign",
            parent=styles["Normal"],
            alignment=2,  # Right align
            fontSize=12,
            textColor=colors.black,
            spaceAfter=10,
        ))

        styles.add(ParagraphStyle(
            name="LeftAlign",
            parent=styles["Normal"],
            alignment=0,  # Left align
            fontSize=10,
            textColor=colors.black,
            spaceAfter=-50,
        ))

        # A custom style for small centered text in the table cells.
        styles.add(ParagraphStyle(
            name="Cnt",
            parent=styles["Normal"],
            alignment=1,
            fontSize=9,
            textColor=colors.black,
        ))
        styles.add(ParagraphStyle(
            name="RightAlignTable",
            parent=styles["Normal"],
            alignment=2,  # Right align
            fontSize=10,
            textColor=colors.black,
            spaceAfter=10,
        ))


        # Story list to hold content
        story = []

        # Create a drawing container with full width (ignoring margins)
        drawing = Drawing(doc.pagesize[0] + doc.leftMargin + doc.rightMargin, 0.1 * inch)
        line = Line(-doc.leftMargin, 0, doc.pagesize[0] + doc.rightMargin, 0)
        drawing.add(line)

        # ===================================================================
        # 6) Seller Info Section
        seller_name = f"{invoice.seller.seller_name}"
        story.append(Paragraph(seller_name, styles["SellerName"]))

        # ===================================================================
        # 4) Add the word "INVOICE" (Centered and styled)
        invoice_title = "INVOICE"
        story.append(Paragraph(invoice_title, styles["InvoiceTitle"]))

        # ===================================================================
        # 5) Invoice Info (Right-aligned - Invoice Number and Date)
        invoice_info = f"""
        <b>invoice number:</b> {invoice.invoice_number}<br/>
        <b>invoice date:</b> {invoice.invoice_date}
        """
        story.append(Paragraph(invoice_info, styles["LeftAlign"]))
        story.append(Spacer(1, 0.5 * inch))

        # Seller Address, Country, Reference
        seller_info = f"""
        <b>Seller name and address:</b><br/>{invoice.seller.seller_name}<br/>{invoice.seller.seller_address}<br/>
        <b> Country of beneficiary:</b> {invoice.seller.seller_country}<br/>
        """

        # ===================================================================
        # 8) Buyer Info Section
        buyer_info = f"""
        <b>Buyer’s Commercial Card No:</b> {invoice.buyer.buyer_card_number}<br/>
        <b> Buyer’s Name:</b> {invoice.buyer.buyer_name}<br/>
        <b> Buyer’s Address:</b> {invoice.buyer.buyer_address}<br/>
        <b> Buyer’s Tel:</b> {invoice.buyer.buyer_tel}

        """
        table_data = [
            [Paragraph(seller_info, styles["Normal"]), Paragraph(buyer_info, styles["Normal"])]
        ]
        
        # Table for Seller and Buyer Info
        info_table = Table(table_data, colWidths=[3.6 * inch, 3.6 * inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'Broadway'),
        ]))
        story.append(Paragraph("Seller and Buyer Info", styles["LeftAlign"]))
        story.append(drawing)
        story.append(info_table)
        story.append(Spacer(0.5, 0.5 * inch))

        # ===================================================================
        # 9) Shipping Info Section (rendered in a two-column table)
        story.append(Paragraph("Shipping Info", styles["LeftAlign"]))
        story.append(drawing)
        story.append(Spacer(0.3, 0.3 * inch))

        shipping_table_data = [
            [
                Paragraph("<b>Transaction Currency:</b> " + invoice.invoice_currency, styles["Normal"]),
                Paragraph("<b>Final delivery place:</b> " + invoice.relevant_location, styles["Normal"])

            ],
            [
                Paragraph("<b>Terms of Delivery:</b> " + invoice.terms_of_delivery, styles["Normal"]),
                Paragraph("<b>Transport mode and means:</b> " + getattr(invoice, 'means_of_transport', ''), styles["Normal"])
            ],
            [
                Paragraph("<b>Country of origin:</b> " + getattr(invoice, 'country_of_origin', ''), styles["Normal"]),
                Paragraph("<b>Port/airport of loading:</b> " + getattr(invoice, 'port_of_loading', ''), styles["Normal"])
            ],
        ]
        shipping_info_table = Table(shipping_table_data, colWidths=[3.5 * inch, 3.5 * inch])
        shipping_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(shipping_info_table)
        story.append(Spacer(1, 1 * inch))
        story.append(Paragraph("Invoice Items", styles["LeftAlign"]))
        story.append(drawing)
        # --- End Shipping Info Table ---

        # ===================================================================
        # 10) Items Table
        table_data = [[
            "No",
            "Item Description",
            "Origin",
            "HS Code",
            "Nw. kg",
            "Gw. kg",
            "Quantity",
            "Unit",
            "Unit Price",
            "Amount"
        ]]

        for idx, item in enumerate(invoice.items.all(), 1):
            description = Paragraph(item.description, styles["Cnt"])
            origin = Paragraph(item.origin, styles["Cnt"])
            commodity_code = Paragraph(str(item.commodity_code), styles["Cnt"])
            nw_kg = str(item.nw_kg)
            gw_kg = str(item.gw_kg)
            unit = Paragraph(str(item.unit), styles["Cnt"])
            quantity = Paragraph(str(item.quantity), styles["Cnt"])
            unit_price = str(item.unit_price)
            try:
                amount_val = float(item.quantity) * float(item.unit_price)
            except (ValueError, TypeError):
                amount_val = 0
            amount = Paragraph(f"{amount_val:.2f}")

            table_data.append([
                str(idx),
                description,
                origin,
                commodity_code,
                nw_kg,
                gw_kg,
                quantity,
                unit,
                unit_price,
                amount
            ])

        col_widths = [
            0.3 * inch, 1.5 * inch, 0.6 * inch, 0.9 * inch,
            0.8 * inch, 0.8 * inch, 0.7 * inch, 0.5 * inch,
            0.9 * inch, 0.9 * inch
        ]
        items_table = Table(table_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(Spacer(1, 0.5 * inch))
        story.append(items_table)
        sub_total_cell = Paragraph("<b>Sub total: </b>" + str(invoice.sub_total) +" "+ invoice.invoice_currency,styles['Cnt'])

        # ===================================================================
        # 11) Cost Summary Table (displayed under the items table)
        total_amount_data = [

            [   "Total", 
                str(invoice.total_nw),
                str(invoice.total_gw),
                str(invoice.total_qty),
                [sub_total_cell]
            ],
        ]
        freight_charge_cell = Paragraph("<b>Freight charge:</b> " + str(invoice.freight_charges)+" "+ invoice.invoice_currency,styles['Cnt'])
        total_amount_cell = Paragraph("<b>Total amount:</b> " + str(invoice.total_amount)+ " "+ invoice.invoice_currency,styles['Cnt'])

        cost_summary_data = [

            [                   
              [freight_charge_cell],

            ],
            [
            [total_amount_cell],

            ]
        ]
        cost_summary_table = Table(total_amount_data, colWidths=[3.3 * inch, 0.8 * inch, 0.8 * inch,0.7 * inch, 2.3 * inch])
        cost_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(cost_summary_table)
        # Set the table style for Total Amount Table
        total_amount_table = Table(cost_summary_data, colWidths=[2.3 * inch])
        total_amount_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        # Set the column widths and adjust the width to make it fit on the right side
        col_widths = [0.8 * inch, 0.7 * inch, 2.4 * inch]  # Adjust the column widths as needed

        # Wrap the table in a container to position it correctly
        table_width = sum(col_widths)  # Get the total width of the table

        # Create a drawing to simulate right-aligning the table
        drawing = Drawing(doc.pagesize[0] - table_width, 0)  # Position the table on the right side
        story.append(drawing)  # Add the drawing to the story
        total_amount_table.hAlign = 'RIGHT'

        # Append the table to the story
        story.append(total_amount_table)

        
   # ===================================================================
        # 12) Banking Info & Seller Seal Table (displayed under cost summary)
        # Left column: Banking info; Right column: Seller seal image (if available)
        banking_info = f"""
        <b>Bank Name:</b> {invoice.seller.seller_bank_name}<br/>
        <b>Account Name:</b> {invoice.seller.seller_account_name}<br/>
        <b>IBAN:</b> {invoice.seller.seller_iban}<br/>
        <b>SWIFT:</b> {invoice.seller.seller_swift}
        """
        right_cell = Paragraph(banking_info, styles["Normal"])

        # Prepare the right cell with seller seal (if available)
        if invoice.seller.seller_seal:
            try:
                seal_image = RLImage(invoice.seller.seller_seal.path, width= 3 *inch, height= 0.75 *inch)
                left_cell = seal_image
            except Exception:
                left_cell = Paragraph("Seal image not available", styles["Normal"])
        else:
            left_cell = Paragraph("No Seal Available", styles["Normal"])

        banking_table_data = [[left_cell, right_cell]]
        banking_info_table = Table(banking_table_data, colWidths=[4 * inch, 4 * inch])
        banking_info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),

        ]))
        story.append(Spacer(1, 0.5 * inch))
        story.append(banking_info_table)

        # ===================================================================
        # 13) Finalize the PDF
        doc.build(story)
        return response
class InvoiceExcelView(APIView):
    """
    Returns an Excel file (XLSX) for the given Invoice (by primary key).
    """
    def get(self, request, pk, format=None):
        try:
            invoice = Invoice.objects.get(pk=pk)
        except Invoice.DoesNotExist:
            return Response({"detail": "Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Write headers
        ws["A1"] = "Invoice ID"
        ws["B1"] = "Invoice Number"
        ws["C1"] = "Seller"
        ws["D1"] = "Buyer"
        ws["E1"] = "Total Amount"
        ws["F1"] = "Freight Charges"
        ws["G1"] = "Currency"
        ws["H1"] = "Invoice Date"

        # Write the invoice info
        ws["A2"] = invoice.invoice_id
        ws["B2"] = invoice.invoice_number
        ws["C2"] = invoice.seller.seller_account_name
        ws["D2"] = invoice.buyer.buyer_card_number
        ws["E2"] = float(invoice.total_amount)
        ws["F2"] = invoice.freight_charges
        ws["G2"] = invoice.invoice_currency
        ws["H2"] = str(invoice.invoice_date)

        # Optionally create a new sheet for items
        item_sheet = wb.create_sheet("Items")
        item_sheet["A1"] = "Description"
        item_sheet["B1"] = "Quantity"
        item_sheet["C1"] = "Unit Price"
        item_sheet["D1"] = "Line Total"

        row = 2
        for item in invoice.items.all():
            item_sheet[f"A{row}"] = item.description
            item_sheet[f"B{row}"] = item.quantity
            item_sheet[f"C{row}"] = float(item.unit_price)
            item_sheet[f"D{row}"] = float(item.line_total)
            row += 1

        # Prepare response with XLSX data
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="invoice_{pk}.xlsx"'

        wb.save(response)
        return response

class PackingPDFView(APIView):
    """
    Returns a PDF with the invoice number and date on the right side,
    and other sections laid out as per the template.
    """
    def get(self, request, pk, format=None):
        try:
            invoice = Invoice.objects.get(pk=pk)
        except Invoice.DoesNotExist:
            return Response({"detail": "Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        # Register the Broadway font
        pdfmetrics.registerFont(TTFont('Broadway', 'static/fonts/broadway.ttf'))

        # 1) Create the HttpResponse for the PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="packing_{pk}.pdf"'

        # Set up the SimpleDocTemplate with custom margins (left, right, top, bottom)
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            title=f"Packing {invoice.invoice_number}",
            author="Your Company",
            leftMargin=0.5 * inch,   # 0.5 inch left margin
            rightMargin=0.5 * inch,  # 0.5 inch right margin
            topMargin=0.2 * inch,    # 0.2 inch top margin
            bottomMargin=0.5 * inch  # 0.5 inch bottom margin
        )

        # 3) Get default styles and add custom styles
        styles = getSampleStyleSheet()

        styles.add(ParagraphStyle(
            name="InvoiceTitle",
            fontSize=24,
            textColor=colors.black,
            alignment=1,  # Center align
            spaceAfter=0
        ))

        styles.add(ParagraphStyle(
            name="SellerName",
            fontName="Broadway",
            fontSize=26,
            textColor=colors.red,
            alignment=1,  # Center align
            spaceAfter=20,
        ))

        styles.add(ParagraphStyle(
            name="RightAlign",
            parent=styles["Normal"],
            alignment=2,  # Right align
            fontSize=12,
            textColor=colors.black,
            spaceAfter=10,
        ))

        styles.add(ParagraphStyle(
            name="LeftAlign",
            parent=styles["Normal"],
            alignment=0,  # Left align
            fontSize=9,
            textColor=colors.black,
            spaceAfter=-50,
        ))

        # A custom style for small centered text in the table cells.
        styles.add(ParagraphStyle(
            name="Cnt",
            parent=styles["Normal"],
            alignment=1,
            fontSize=9,
            textColor=colors.black,
        ))

        # Story list to hold content
        story = []

        # Create a drawing container with full width (ignoring margins)
        drawing = Drawing(doc.pagesize[0] + doc.leftMargin + doc.rightMargin, 0.1 * inch)
        line = Line(-doc.leftMargin, 0, doc.pagesize[0] + doc.rightMargin, 0)
        drawing.add(line)

        # ===================================================================
        # 6) Seller Info Section
        seller_name = f"{invoice.seller.seller_name}"
        story.append(Paragraph(seller_name, styles["SellerName"]))

        # ===================================================================
        # 4) Add the word "INVOICE" (Centered and styled)
        invoice_title = "PACKING"
        story.append(Paragraph(invoice_title, styles["InvoiceTitle"]))

        # ===================================================================
        # 5) Invoice Info (Right-aligned - Invoice Number and Date)
        invoice_info = f"""
        <b>packing list number: </b>{invoice.invoice_number}<br/>
        <b>packing list date: </b>{invoice.invoice_date}
        """
        story.append(Paragraph(invoice_info, styles["LeftAlign"]))
        story.append(Spacer(1, 0.5 * inch))

        # Seller Address, Country, Reference
        seller_info = f"""
        <b>Seller name and address:</b><br/>{invoice.seller.seller_name}<br/>{invoice.seller.seller_address}<br/>
        <b> Country of beneficiary:</b> {invoice.seller.seller_country}<br/>
        """

        # ===================================================================
        # 8) Buyer Info Section
        buyer_info = f"""
        <b>Buyer’s Commercial Card No:</b> {invoice.buyer.buyer_card_number}<br/>
        <b> Buyer’s Name:</b> {invoice.buyer.buyer_name}<br/>
        <b> Buyer’s Address:</b> {invoice.buyer.buyer_address}<br/>
        <b> Buyer’s Tel:</b> {invoice.buyer.buyer_tel}

        """
        table_data = [
            [Paragraph(seller_info, styles["Normal"]), Paragraph(buyer_info, styles["Normal"])]
        ]
        
        # Table for Seller and Buyer Info
        info_table = Table(table_data, colWidths=[3.6 * inch, 3.6 * inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'Broadway'),
        ]))
        story.append(Paragraph("Seller and Buyer Info", styles["LeftAlign"]))
        story.append(drawing)
        story.append(info_table)
        story.append(Spacer(0.5, 0.5 * inch))

        # ===================================================================
        # 9) Shipping Info Section (rendered in a two-column table)
        story.append(Paragraph("Shipping Info", styles["LeftAlign"]))
        story.append(drawing)

        shipping_table_data = [
            [
                Paragraph("<b>Transaction Currency:</b> " + invoice.invoice_currency, styles["Normal"]),
                Paragraph("<b>Final delivery place:</b> " + invoice.relevant_location, styles["Normal"])

            ],
            [
                Paragraph("<b>Terms of Delivery:</b> " + invoice.terms_of_delivery, styles["Normal"]),
                Paragraph("<b>Transport mode and means:</b> " + getattr(invoice, 'means_of_transport', ''), styles["Normal"])
            ],
            [
                Paragraph("<b>Country of origin:</b> " + getattr(invoice, 'country_of_origin', ''), styles["Normal"]),
                Paragraph("<b>Port/airport of loading:</b> " + getattr(invoice, 'port_of_loading', ''), styles["Normal"])
            ],
        ]
        shipping_info_table = Table(shipping_table_data, colWidths=[3.5 * inch, 3.5 * inch])
        shipping_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(shipping_info_table)
        story.append(Spacer(0.5, 0.5 * inch))
        story.append(Paragraph("Invoice Items", styles["LeftAlign"]))
        story.append(drawing)
        # --- End Shipping Info Table ---

        # ===================================================================
        # 10) Items Table
        table_data = [[
            "No",
            "Item Description",
            "Origin",
            "HS Code",
            "Nw. kg",
            "Gw. kg",
            "Quantity",           
            "Unit",
            "PCK",

        ]]

        for idx, item in enumerate(invoice.items.all(), 1):
            description = Paragraph(item.description, styles["Cnt"])
            origin = Paragraph(item.origin, styles["Cnt"])
            commodity_code = Paragraph(str(item.commodity_code), styles["Cnt"])
            nw_kg = str(item.nw_kg)
            gw_kg = str(item.gw_kg)
            unit = Paragraph(str(item.unit), styles["Cnt"])
            quantity = str(item.quantity)
            pack = str(item.pack)


            table_data.append([
                str(idx),
                description,
                origin,
                commodity_code,
                nw_kg,
                gw_kg,
                quantity,
                unit,
                pack,

            ])

        col_widths = [
            0.3 * inch, 1.5 * inch, 0.7 * inch, 0.9 * inch,
            0.8 * inch, 0.8 * inch, 0.7 * inch, 0.5 * inch,
            0.9 * inch, 1 * inch
        ]
        items_table = Table(table_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(Spacer(1, 0.5 * inch))
        story.append(items_table)

        # ===================================================================
               # 11) Cost Summary Table (displayed under the items table)
        # Calculate subtotal as (total_amount - freight_charges)
        subtotal = invoice.total_amount - invoice.freight_charges
        cost_summary_data = [

            [   "Total", 
                str(invoice.total_nw),
                str(invoice.total_gw),
                str(invoice.total_qty),
                str(invoice.total_pack)

            ]
        ]
        cost_summary_table = Table(cost_summary_data, colWidths=[3.4 * inch, 0.8 * inch, 0.8 * inch, 0.7 * inch, 1.4 * inch])
        cost_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(cost_summary_table)

     # ===================================================================
        # 12) Banking Info & Seller Seal Table (displayed under cost summary)
        # Left column: Banking info; Right column: Seller seal image (if available)
        banking_info = f"""
        <b>Bank Name:</b> {invoice.seller.seller_bank_name}<br/>
        <b>Account Name:</b> {invoice.seller.seller_account_name}<br/>
        <b>IBAN:</b> {invoice.seller.seller_iban}<br/>
        <b>SWIFT:</b> {invoice.seller.seller_swift}
        """
        right_cell = Paragraph(banking_info, styles["Normal"])

        # Prepare the right cell with seller seal (if available)
        if invoice.seller.seller_seal:
            try:
                seal_image = RLImage(invoice.seller.seller_seal.path, width= 3 *inch, height= 0.75 *inch)
                left_cell = seal_image
            except Exception:
                left_cell = Paragraph("Seal image not available", styles["Normal"])
        else:
            left_cell = Paragraph("No Seal Available", styles["Normal"])

        banking_table_data = [[left_cell, right_cell]]
        banking_info_table = Table(banking_table_data, colWidths=[4 * inch, 4 * inch])
        banking_info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),

        ]))
        story.append(Spacer(1, 0.5 * inch))
        story.append(banking_info_table)

        # ===================================================================
        # 13) Finalize the PDF
        doc.build(story)
        return response

class CertiOriginView(APIView):
    def post(self, request):
        # 1) Fetch the template based on template_id
        try:
            template_id = request.data.get('template_id')
            template = ImageTemplate.objects.get(id=template_id)
        except ImageTemplate.DoesNotExist:
            return Response({"error": "Template not found."},
                            status=status.HTTP_404_NOT_FOUND)

        # 2) Validate incoming data
        serializer = OriginSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors,
                            status=status.HTTP_400_BAD_REQUEST)

        # 3) Extract validated fields
        exporter = serializer.validated_data['exporter']
        consignee = serializer.validated_data['consignee']
        means_of_transport = serializer.validated_data['means_of_transport']
        goods = serializer.validated_data['goods']  # List of dictionaries

        # 4) Load the template image and prepare for drawing
        img = Image.open(template.template_image.path)
        draw = ImageDraw.Draw(img)
        font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'arial.ttf')
        font_size = template.font_size                # <--- Read from the model
        font = ImageFont.truetype(font_path, font_size)

        # 5) Draw static fields using multiline_text to preserve line breaks
        # Positions are stored as "x,y" strings in your model.
        exporter_x, exporter_y = map(int, template.exporter_position.split(','))
        consignee_x, consignee_y = map(int, template.consignee_position.split(','))
        transport_x, transport_y = map(int, template.means_of_transport_position.split(','))

        draw.multiline_text((exporter_x, exporter_y), exporter, fill="black", font=font)
        draw.multiline_text((consignee_x, consignee_y), consignee, fill="black", font=font)
        draw.multiline_text((transport_x, transport_y), means_of_transport, fill="black", font=font)

        # 6) Get the X-axis positions for each goods column.
        #    In your model, these fields hold a single integer (as a string).
        #    Additionally, we define an X coordinate for the index column.
        description_x = int(template.description_position)       # e.g., "100"
        hscode_x = int(template.hscode_position)                 # e.g., "300"
        quantity_x = int(template.quantity_position)             # e.g., "500"
        invoices_x = int(template.number_of_invoices_position)     # e.g., "700"

        # 7) Get the start position and line height for listing goods.
        #    goods_start_position is expected to be a string like "100,300" (x,y).
        goods_start_x, goods_start_y = map(int, template.goods_start_position.split(','))
        current_y = goods_start_y
        line_height = template.goods_line_height  # Vertical spacing between each row
        index_x = goods_start_x  
        # 8) Draw each good in its own row.
        #    Include an index column (using good['index'] if provided, or falling back to the loop index).
        for i, good in enumerate(goods, start=1):
            index_value = good.get('index', i)
            description_text = good.get('description', '')
            hscode_text = good.get('hscode', '')
            quantity_text = good.get('quantity', '')
            invoices_text = good.get('number_of_invoices', '')

            # Draw each field at its respective X position and the current Y position.
            draw.multiline_text((index_x,       current_y), str(index_value), fill="black", font=font)
            draw.multiline_text((description_x, current_y), description_text, fill="black", font=font)
            draw.multiline_text((hscode_x,      current_y), hscode_text, fill="black", font=font)
            draw.multiline_text((quantity_x,    current_y), quantity_text, fill="black", font=font)
            draw.multiline_text((invoices_x,    current_y), invoices_text, fill="black", font=font)

            # Move down to the next row.
            current_y += line_height

        # 9) Save the modified image to a BytesIO object and return it.
        img_io = io.BytesIO()
        img.save(img_io, format="JPEG")
        img_io.seek(0)

        return HttpResponse(img_io.getvalue(), content_type="image/jpeg")
class TemplateListView(APIView):    
    def get(self, request):
        # Get all templates from the database
        templates = ImageTemplate.objects.all()
        serializer = ImageTemplateSerializer(templates, many=True)
        return Response(serializer.data)


class ProformaInvoicePDFView(APIView):

    """
    Returns a PDF with the invoice number and date on the right side,
    and other sections laid out as per the template.
    """
    def get(self, request, pk, format=None):
        try:
            invoice = ProformaInvoice.objects.get(pk=pk)
        except ProformaInvoice.DoesNotExist:
            return Response({"detail": "Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        # Register the Broadway font
        pdfmetrics.registerFont(TTFont('Broadway', 'static/fonts/broadway.ttf'))

        # 1) Create the HttpResponse for the PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="proforma_{pk}.pdf"'

        # Set up the SimpleDocTemplate with custom margins (left, right, top, bottom)
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            title=f"Proforma {invoice.proforma_invoice_number}",
            author="Your Company",
            leftMargin=0.1 * inch,   # 0.5 inch left margin
            rightMargin=0.1 * inch,  # 0.5 inch right margin
            topMargin=0.2 * inch,    # 0.2 inch top margin
            bottomMargin=0.5 * inch  # 0.5 inch bottom margin
        )

        # 3) Get default styles and add custom styles
        styles = getSampleStyleSheet()

        styles.add(ParagraphStyle(
            name="InvoiceTitle",
            fontSize=24,
            textColor=colors.black,
            alignment=1,  # Center align
            spaceAfter=0
        ))

        styles.add(ParagraphStyle(
            name="SellerName",
            fontName="Broadway",
            fontSize=26,
            textColor=colors.red,
            alignment=1,  # Center align
            spaceAfter=20,
        ))

        styles.add(ParagraphStyle(
            name="RightAlign",
            parent=styles["Normal"],
            alignment=2,  # Right align
            fontSize=12,
            textColor=colors.black,
            spaceAfter=10,
        ))

        styles.add(ParagraphStyle(
            name="LeftAlign",
            parent=styles["Normal"],
            alignment=0,  # Left align
            fontSize=10,
            textColor=colors.black,
            spaceAfter=-50,
        ))

        # A custom style for small centered text in the table cells.
        styles.add(ParagraphStyle(
            name="Cnt",
            parent=styles["Normal"],
            alignment=1,
            fontSize=9,
            textColor=colors.black,
        ))
        styles.add(ParagraphStyle(
            name="RightAlignTable",
            parent=styles["Normal"],
            alignment=2,  # Right align
            fontSize=10,
            textColor=colors.black,
            spaceAfter=10,
        ))


        # Story list to hold content
        story = []

        # Create a drawing container with full width (ignoring margins)
        drawing = Drawing(doc.pagesize[0] + doc.leftMargin + doc.rightMargin, 0.1 * inch)
        line = Line(-doc.leftMargin, 0, doc.pagesize[0] + doc.rightMargin, 0)
        drawing.add(line)

        # ===================================================================
        # 6) Seller Info Section
        seller_name = f"{invoice.seller.seller_name}"
        story.append(Paragraph(seller_name, styles["SellerName"]))

        # ===================================================================
        # 4) Add the word "INVOICE" (Centered and styled)
        invoice_title = "PROFROMA"
        story.append(Paragraph(invoice_title, styles["InvoiceTitle"]))

        # ===================================================================
        # 5) Invoice Info (Right-aligned - Invoice Number and Date)
        invoice_info = f"""
        <b>proforma number:</b> {invoice.proforma_invoice_number}<br/>
        <b>proforma date:</b> {invoice.proforma_invoice_date}
        """
        story.append(Paragraph(invoice_info, styles["LeftAlign"]))
        story.append(Spacer(1, 0.5 * inch))

        # Seller Address, Country, Reference
        seller_info = f"""
        <b>Seller name and address:</b><br/>{invoice.seller.seller_name}<br/>{invoice.seller.seller_address}<br/>
        <b> Country of beneficiary:</b> {invoice.seller.seller_country}<br/>
        """

        # ===================================================================
        # 8) Buyer Info Section
        buyer_info = f"""
        <b>Buyer’s Commercial Card No:</b> {invoice.buyer.buyer_card_number}<br/>
        <b> Buyer’s Name:</b> {invoice.buyer.buyer_name}<br/>
        <b> Buyer’s Address:</b> {invoice.buyer.buyer_address}<br/>
        <b> Buyer’s Tel:</b> {invoice.buyer.buyer_tel}

        """
        table_data = [
            [Paragraph(seller_info, styles["Normal"]), Paragraph(buyer_info, styles["Normal"])]
        ]
        
        # Table for Seller and Buyer Info
        info_table = Table(table_data, colWidths=[3.6 * inch, 3.6 * inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'Broadway'),
        ]))
        story.append(Paragraph("Seller and Buyer Info", styles["LeftAlign"]))
        story.append(drawing)
        story.append(info_table)
        story.append(Spacer(0.5, 0.5 * inch))

        # ===================================================================
        # 9) Shipping Info Section (rendered in a two-column table)
        story.append(Paragraph("Shipping Info", styles["LeftAlign"]))
        story.append(drawing)
        story.append(Spacer(0.3, 0.3 * inch))

        shipping_table_data = [
            [
                Paragraph("<b>Transaction Currency:</b> " + invoice.proforma_invoice_currency, styles["Normal"]),
                Paragraph("<b>Final delivery place:</b> " + invoice.relevant_location, styles["Normal"])

            ],
            [
                Paragraph("<b>Terms of Delivery:</b> " + invoice.terms_of_delivery, styles["Normal"]),
                Paragraph("<b>Transport mode and means:</b> " + getattr(invoice, 'means_of_transport', ''), styles["Normal"])
            ],
            [
                Paragraph("<b>Country of origin:</b> " + getattr(invoice, 'country_of_origin', ''), styles["Normal"]),
                Paragraph("<b>Port/airport of loading:</b> " + getattr(invoice, 'port_of_loading', ''), styles["Normal"])
            ],
            [
                Paragraph("<b>Terms of Payment:</b> " + getattr(invoice, 'terms_of_payment', ''), styles["Normal"]),
                Paragraph("<b>Partial Shipment:</b> " + ("Allowed" if getattr(invoice, 'partial_shipment', False) else "Not Allowed"), styles["Normal"])
            ],            [
                Paragraph("<b>Standard:</b> " + getattr(invoice, 'standard', ''), styles["Normal"]),

            ],
        ]
        shipping_info_table = Table(shipping_table_data, colWidths=[3.5 * inch, 3.5 * inch])
        shipping_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        story.append(shipping_info_table)
        story.append(Spacer(1, 1 * inch))
        story.append(Paragraph("Invoice Items", styles["LeftAlign"]))
        story.append(drawing)
        # --- End Shipping Info Table ---

        # ===================================================================
        # 10) Items Table
        table_data = [[
            "No",
            "Item Description",
            "Origin",
            "HS Code",
            "Nw. kg",
            "Gw. kg",
            "Quantity",
            "Unit",
            "Unit Price",
            "Amount"
        ]]

        for idx, item in enumerate(invoice.items.all(), 1):
            description = Paragraph(item.description, styles["Cnt"])
            origin = Paragraph(item.origin, styles["Cnt"])
            commodity_code = Paragraph(str(item.commodity_code), styles["Cnt"])
            nw_kg = str(item.nw_kg)
            gw_kg = str(item.gw_kg)
            unit = Paragraph(str(item.unit), styles["Cnt"])
            quantity = str(item.quantity)
            unit_price = str(item.unit_price)
            try:
                amount_val = float(item.quantity) * float(item.unit_price)
            except (ValueError, TypeError):
                amount_val = 0
            amount = f"{amount_val:.2f}"

            table_data.append([
                str(idx),
                description,
                origin,
                commodity_code,
                nw_kg,
                gw_kg,
                quantity,
                unit,
                unit_price,
                amount
            ])

        col_widths = [
            0.3 * inch, 1.5 * inch, 0.6 * inch, 0.9 * inch,
            0.8 * inch, 0.8 * inch, 0.7 * inch, 0.5 * inch,
            0.9 * inch, 0.9 * inch
        ]
        items_table = Table(table_data, colWidths=col_widths)
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(Spacer(1, 0.5 * inch))
        story.append(items_table)
        sub_total_cell = Paragraph("<b>Sub total: </b>" + str(invoice.sub_total) +" "+ invoice.proforma_invoice_currency,styles['Cnt'])

        # ===================================================================
        # 11) Cost Summary Table (displayed under the items table)
        total_amount_data = [

            [   "Total", 
                str(invoice.total_nw),
                str(invoice.total_gw),
                str(invoice.total_qty),
                [sub_total_cell]
            ],
        ]
        freight_charge_cell = Paragraph("<b>Freight charge:</b> " + str(invoice.proforma_freight_charges)+" "+ invoice.proforma_invoice_currency,styles['Cnt'])
        total_amount_cell = Paragraph("<b>Total amount:</b> " + str(invoice.total_amount)+ " "+ invoice.proforma_invoice_currency,styles['Cnt'])

        cost_summary_data = [

            [                   
              [freight_charge_cell],

            ],
            [
            [total_amount_cell],

            ]
        ]
        cost_summary_table = Table(total_amount_data, colWidths=[3.3 * inch, 0.8 * inch, 0.8 * inch,0.7 * inch, 2.3 * inch])
        cost_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(cost_summary_table)
        # Set the table style for Total Amount Table
        total_amount_table = Table(cost_summary_data, colWidths=[2.3 * inch])
        total_amount_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        # Set the column widths and adjust the width to make it fit on the right side
        col_widths = [0.8 * inch, 0.7 * inch, 2.4 * inch]  # Adjust the column widths as needed

        # Wrap the table in a container to position it correctly
        table_width = sum(col_widths)  # Get the total width of the table

        # Create a drawing to simulate right-aligning the table
        drawing = Drawing(doc.pagesize[0] - table_width, 0)  # Position the table on the right side
        story.append(drawing)  # Add the drawing to the story
        total_amount_table.hAlign = 'RIGHT'

        # Append the table to the story
        story.append(total_amount_table)

        
   # ===================================================================
        # 12) Banking Info & Seller Seal Table (displayed under cost summary)
        # Left column: Banking info; Right column: Seller seal image (if available)
        banking_info = f"""
        <b>Bank Name:</b> {invoice.seller.seller_bank_name}<br/>
        <b>Account Name:</b> {invoice.seller.seller_account_name}<br/>
        <b>IBAN:</b> {invoice.seller.seller_iban}<br/>
        <b>SWIFT:</b> {invoice.seller.seller_swift}
        """
        right_cell = Paragraph(banking_info, styles["Normal"])

        # Prepare the right cell with seller seal (if available)
        if invoice.seller.seller_seal:
            try:
                seal_image = RLImage(invoice.seller.seller_seal.path, width= 3 *inch, height= 0.75 *inch)
                left_cell = seal_image
            except Exception:
                left_cell = Paragraph("Seal image not available", styles["Normal"])
        else:
            left_cell = Paragraph("No Seal Available", styles["Normal"])

        banking_table_data = [[left_cell, right_cell]]
        banking_info_table = Table(banking_table_data, colWidths=[4 * inch, 4 * inch])
        banking_info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),

        ]))
        story.append(Spacer(1, 0.5 * inch))
        story.append(banking_info_table)

        # ===================================================================
        # 13) Finalize the PDF
        doc.build(story)
        return response


class CombinedPDFView(APIView):
    def get(self, request, pk, format=None):
        # Extract the template_id from the request data
        template_id = request.query_params.get('template_id')
        if not template_id:
            return Response({"error": "Template ID is required."}, status=400)

        # Initialize PDF merger
        merger = PdfMerger()

        # Create a factory to simulate a POST request for OverlayTextView
        factory = APIRequestFactory()
        overlay_text_request = factory.post(
            '/overlay/', 
            data={'template_id': template_id, 'invoice_id': pk},  # Passing template_id and invoice_id in the payload
            format='json'
        )
        overlay_text_request.META['HTTP_AUTHORIZATION'] = request.headers.get('Authorization')

        # Overlay Text PDF generation (using OverlayTextView)
        overlay_text_view = OverlayTextView.as_view()
        overlay_text_response = overlay_text_view(overlay_text_request)
        
        # Check if the response content is a valid PDF
        if self.is_valid_pdf(overlay_text_response.content):
            overlay_text_pdf = BytesIO(overlay_text_response.content)
            merger.append(overlay_text_pdf)
        else:
            return Response({"error": "Invalid PDF generated for Overlay Text"}, status=400)

        # Invoice PDF generation (using InvoicePDFView)
        invoice_http_request = factory.get(f'/api/documents/invoice/{pk}/pdf/', format='pdf')
        invoice_http_request.META['HTTP_AUTHORIZATION'] = request.headers.get('Authorization')
        
        invoice_pdf_view = InvoicePDFView.as_view()
        invoice_pdf_response = invoice_pdf_view(invoice_http_request, pk)

        if self.is_valid_pdf(invoice_pdf_response.content):
            invoice_pdf = BytesIO(invoice_pdf_response.content)
            merger.append(invoice_pdf)
        else:
            return Response({"error": "Invalid PDF generated for Invoice"}, status=400)

        # Packing PDF generation (using PackingPDFView)
        packing_http_request = factory.get(f'/api/documents/packing/{pk}/pdf/', format='pdf')
        packing_http_request.META['HTTP_AUTHORIZATION'] = request.headers.get('Authorization')
        
        packing_pdf_view = PackingPDFView.as_view()
        packing_pdf_response = packing_pdf_view(packing_http_request, pk)

        if self.is_valid_pdf(packing_pdf_response.content):
            packing_pdf = BytesIO(packing_pdf_response.content)
            merger.append(packing_pdf)
        else:
            return Response({"error": "Invalid PDF generated for Packing"}, status=400)

        # Create a final combined PDF
        final_pdf = BytesIO()
        merger.write(final_pdf)
        final_pdf.seek(0)

        # Prepare the HttpResponse to serve the combined PDF
        response = HttpResponse(final_pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="combined_invoice.pdf"'
        
        return response

    def is_valid_pdf(self, content):
        """Helper method to check if the content is a valid PDF."""
        try:
            # Try reading the content as a PDF
            pdf = BytesIO(content)
            pdf.seek(0)
            merger = PdfMerger()
            merger.append(pdf)
            return True
        except Exception as e:
            print(f"PDF validation error: {e}")
            return False