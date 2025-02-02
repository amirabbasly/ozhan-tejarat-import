from rest_framework.viewsets import ModelViewSet
from .models import Representation, Check
from .serializers import RepresentationSerializer, CheckSerializer, SummarySerializer
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import  status
from openpyxl import load_workbook
from django.db.models import Sum, Count
from datetime import datetime
from .utils import jalali_to_gregorian  # You need a utility to convert Jalali to Gregorian


class RepresentationViewSet(ModelViewSet):
    queryset = Representation.objects.all()
    serializer_class = RepresentationSerializer
    parser_classes = (MultiPartParser, FormParser)

class CheckViewSet(ModelViewSet):
    queryset = Check.objects.all()
    serializer_class = CheckSerializer
    
class ImportRepresentationsView(APIView):
    """
    DRF view to handle Excel file uploads with Persian headings.
    """
    parser_classes = (MultiPartParser, FormParser,)  # allow file uploads

    EXPECTED_HEADINGS = {
        'موکل': 'representi',
        'نام وکیل': 'representor',
        'درخواست دهنده': 'applicant',
        'تاریخ شروع': 'start_date',
        'تاریخ انقضا': 'end_date',
        'توکیل به غیر': 'another_deligation',
        'عزل وکیل': 'representor_dismissal',
        'خلاصه وکالت': 'representation_summary',
        'شناسه سند': 'doc_number',
        'رمز تصدیق': 'verification_code',

    }

    def post(self, request, format=None):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active  # or workbook["SheetName"] if needed

            # 1) Validate heading row
            heading_row = [cell.value for cell in sheet[1]]  # First row as list
            if len(heading_row) < len(self.EXPECTED_HEADINGS):
                return Response({"error": "تعداد ستون های فایل کامل نمیباشد"}, status=400)

            # Build a map from column index -> model field name
            col_map = {}
            for col_index, heading in enumerate(heading_row):
                if heading in self.EXPECTED_HEADINGS:
                    col_map[col_index] = self.EXPECTED_HEADINGS[heading]

            # Check we found all required headings
            if len(col_map) < len(self.EXPECTED_HEADINGS):
                missing_headings = set(self.EXPECTED_HEADINGS.keys()) - set(heading_row)
                return Response({
                    "error": f"فایل باید شامل فیلد های: {', '.join(missing_headings)} باشد"
                }, status=400)

            # 2) Iterate from row 2 onward
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows (all None)
                if all(value is None for value in row):
                    continue

                data_for_model = {}

                for col_index, cell_value in enumerate(row):
                    if col_index in col_map:
                        field_name = col_map[col_index]

                        # Handle boolean fields
                        if field_name in ['another_deligation', 'representor_dismissal']:
                            if cell_value in [None, "ندارد"]:
                                data_for_model[field_name] = False
                            elif cell_value in ["دارد", "بله"]:
                                data_for_model[field_name] = True
                            else:
                                data_for_model[field_name] = False  # Default to False for any other unexpected input
                        else:
                            data_for_model[field_name] = cell_value

                doc_number = data_for_model.get('doc_number')
                if not doc_number:
                    continue

                Representation.objects.update_or_create(
                    doc_number=doc_number,
                    defaults=data_for_model
                )

            return Response({"success": "وکالتنامه ها با موفقیت بارگزاری شدند."})

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ImportChecksView(APIView):
    """
    DRF view to handle Excel file uploads for the Check model with Persian headings.
    """
    parser_classes = (MultiPartParser, FormParser,)  # allow file uploads

    EXPECTED_HEADINGS = {
        'صادر کننده': 'issuer',
        'کد چک': 'check_code',
        'تاریخ': 'date',
        'مبلغ': 'value',
        'در وجه': 'issued_for',
        'بانک': 'bank',
        'پاس شده': 'is_paid',
    }

    def post(self, request, format=None):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active  # or workbook["SheetName"] if needed

            # 1) Validate heading row
            heading_row = [cell.value for cell in sheet[1]]  # First row as list
            if len(heading_row) < len(self.EXPECTED_HEADINGS):
                return Response({"error": "تعداد ستون های فایل کامل نمیباشد"}, status=400)

            # Build a map from column index -> model field name
            col_map = {}
            for col_index, heading in enumerate(heading_row):
                if heading in self.EXPECTED_HEADINGS:
                    col_map[col_index] = self.EXPECTED_HEADINGS[heading]

            # Check we found all required headings
            if len(col_map) < len(self.EXPECTED_HEADINGS):
                missing_headings = set(self.EXPECTED_HEADINGS.keys()) - set(heading_row)
                return Response({
                    "error": f"فایل باید شامل فیلد های: {', '.join(missing_headings)} باشد"
                }, status=400)

            # 2) Iterate from row 2 onward
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip empty rows (all None)
                if all(value is None for value in row):
                    continue

                data_for_model = {}

                for col_index, cell_value in enumerate(row):
                    if col_index in col_map:
                        field_name = col_map[col_index]

                        # Handle boolean field (is_paid)
                        if field_name == 'is_paid':
                            if cell_value in [None, "ندارد"]:
                                data_for_model[field_name] = False
                            elif cell_value in ["دارد", "بله"]:
                                data_for_model[field_name] = True
                            else:
                                data_for_model[field_name] = False  # Default to False for any other unexpected input
                        else:
                            data_for_model[field_name] = cell_value

                check_code = data_for_model.get('check_code')
                if not check_code:
                    continue

                Check.objects.update_or_create(
                    check_code=check_code,
                    defaults=data_for_model
                )

            return Response({"success": "چک‌ها با موفقیت بارگزاری شدند."})

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class SummaryView(APIView):
    
    def get(self, request):
        # Get the number of passed and unpassed checks
        passed_checks = Check.objects.filter(is_paid=True)
        unpassed_checks = Check.objects.filter(is_paid=False)

        passed_checks_count = passed_checks.count()
        unpassed_checks_count = unpassed_checks.count()

        # Get the total value of passed and unpassed checks
        passed_checks_value = passed_checks.aggregate(total_value=Sum('value'))['total_value'] or 0
        unpassed_checks_value = unpassed_checks.aggregate(total_value=Sum('value'))['total_value'] or 0

        # Get the number of representations where the end date has passed
        today = datetime.today().date()  # Today's Gregorian date
        past_representations_count = 0
        unpast_representations_count = 0

        # Iterate through representations and compare their end_date (Jalali) with today's Gregorian date
        representations = Representation.objects.all()
        for rep in representations:
            # Convert end_date (Jalali) to Gregorian using the utility function
            end_date_gregorian = jalali_to_gregorian(rep.end_date)
            
            if end_date_gregorian and end_date_gregorian <= today:
                past_representations_count += 1
            else:
                unpast_representations_count += 1

        # Prepare the summary data
        summary_data = {
            'passed_checks_count': passed_checks_count,
            'unpassed_checks_count': unpassed_checks_count,
            'passed_checks_value': passed_checks_value,
            'unpassed_checks_value': unpassed_checks_value,
            'past_representations_count': past_representations_count,
            'unpast_representations_count': unpast_representations_count,
        }

        # Serialize and return the response
        serializer = SummarySerializer(summary_data)
        return Response(serializer.data)